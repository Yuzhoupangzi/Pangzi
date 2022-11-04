import openpyxl as op
import pandas as pd
import os
df = pd.read_excel('验收表.xlsx')

data = pd.DataFrame(df)

contract_number = data['CMS合同号']
contract_name = data['合同名称']
contract_money = data['合同金额']
contract_father = data['甲方']
contract_keyperson = data['百胜\n负责人']
contract_inspection_rating = data['本次验收比']

for i in range(0,len(list(set(contract_keyperson)))):
    if os.path.exists(list(set(contract_keyperson))[i]) == 0:
        os.mkdir(list(set(contract_keyperson))[i])


wb = op.load_workbook('模板.xlsx')
for i in range(0,len(contract_number)):
    ws = wb['Sheet1']
    excel_name = contract_keyperson[i].replace('\n','') + '-' + contract_number[i] + '-' + contract_name[i].replace('/','-') + '.xlsx'
    money = round(contract_money[i] * contract_inspection_rating[i] * 1000,2)
    contract_inspection_rate = "%.0f%%" % (contract_inspection_rating[i] * 100)
    text = f'经项目双方的共同努力，本项目已经在合同期内完成{contract_inspection_rate}，甲方组织验收，已达到甲方的要求。甲方给予本项目验收通过的结论。本次验收总金额为：{money}元。'
    ws.cell(7,1,contract_father[i])
    ws.cell(7,2,contract_number[i])
    ws.cell(7,3,contract_name[i])
    ws.cell(7,5,money)
    ws.cell(9,1,text)
    wb.save('./'+ contract_keyperson[i] + '/' + excel_name)
    print(excel_name + '已经创建完毕')

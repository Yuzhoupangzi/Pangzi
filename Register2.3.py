# coding=gbk
import xlrd
from xlwt import *
import pandas as pd
import openpyxl as op
import datetime

add_list=[]
add_list_new=[]

dic1 ={'1': '摄像头项目对账清单明细.xlsx',
       '2': '无人收货项目对账清单明细.xlsx',
       '3': '瓦力项目对账清单明细.xlsx',
       '4': '手表一代项目对账清单明细.xlsx',
       '5': '手表二代项目对账清单明细.xlsx',
       '6': 'RFID智能货物管理项目对账清单明细.xlsx',
       '7': 'PHAI项目对账清单明细.xlsx',
       '8': 'ELO叫号屏项目对账清单明细.xlsx',
       '9': 'IOT项目对账清单明细.xlsx',
       '10': 'IOT-485项目对账清单明细.xlsx',
       }

fileName='厂商订单汇总明细.xlsx'
bk=xlrd.open_workbook(fileName)
shxrange=range(bk.nsheets)

try:
  sh=bk.sheet_by_name("Sheet1")
except:
  print("代码出错")
nrows=sh.nrows #获取行数
book = Workbook(encoding='utf-8')
sheet = book.add_sheet('Sheet1') #创建一个sheet
for i in range(0,nrows):
  row_data=sh.row_values(i)
  #获取第i行第3列数据
  #sh.cell_value(i,3)
  #---------写出文件到excel--------

  # print ("写入 "+str(i)+" 行")
  sheet.write(i,0, label = sh.cell_value(i,4))
  sheet.write(i,1, label = sh.cell_value(i,23)) #向第1行第1列写入获取到的值
  sheet.write(i,2, label = sh.cell_value(i,24))  #向第1行第2列写入获取到的值
  sheet.write(i,3, label = sh.cell_value(i,15))
  sheet.write(i,4, label = sh.cell_value(i,34))
  sheet.write(i,5, label = sh.cell_value(i,35))
  sheet.write(i,6, label = sh.cell_value(i,5))
  sheet.write(i,7, label = sh.cell_value(i,1))
  sheet.write(i,8, label=sh.cell_value(i,12))
  sheet.write(i,9, label=sh.cell_value(i,36))
  sheet.write(i,10, label=sh.cell_value(i,37))
  sheet.write(i,11, label=sh.cell_value(i,39))
book.save('Bridge.xls')


shijian = datetime.date.today()
print(shijian)
print('*************************')
print("1:摄像头")
print("2:无人收货")
print('3:瓦力')
print('4:手表一代')
print('5:手表二代')
print('6:RFID智能货物管理')
print('7:PHAI')
print('8:ELO叫号屏')
print('9:IOT')
print('10:IOT-485')
print('*************************')
num = input('请输入要登记台账名称：')

add_Name_list5 = []
add_Market = []
add_Name_list5 = []
add_Market = []
add_Name_list = []
add_Name_list2 = []
add_Name_list4 = []
add_Product_id_list = []
add_Product_name_num = []
add_Product_name_list = []
add_Name_list6 = []
add_Name_list3 = []
add_Order_num = []
add_Price = []

add_Name_list5_new = []
add_Market_new = []
add_Name_list5_new = []
add_Market_new = []
add_Name_list_new = []
add_Name_list2_new = []
add_Name_list4_new = []
add_Product_id_list_new = []
add_Product_name_num_new = []
add_Product_name_list_new = []
add_Name_list6_new = []
add_Name_list3_new = []
add_Order_num_new = []
add_Price_new = []



if dic1[num] == '摄像头项目对账清单明细.xlsx':
    df = pd.read_excel('摄像头项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('摄像头项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == '海康AI摄像头DS2XA2346FMLS-1台' :
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('摄像头项目对账清单明细.xlsx')

    write()

if dic1[num] == '无人收货项目对账清单明细.xlsx':
    df = pd.read_excel('无人收货项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('无人收货项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == '双向双扇玻璃门智能门锁-1套' or Product_name_list[a] == '双向单扇玻璃门智能门锁-1套' or Product_name_list[a] == '双向双扇玻璃门智能门锁金色版-1套' or Product_name_list[a] == '双向单扇玻璃门智能门锁金色版-1套' or Product_name_list[a] == '单向单扇木质门智能门锁-1套' or Product_name_list[a] == '防水磁力锁支架-1台' or Product_name_list[a] == '电插锁U型上下支架-1台' or Product_name_list[a] == '单向单扇木质门智能门锁金色版-1套'  or Product_name_list[a] == '电插锁U型下支架-1台' or Product_name_list[a] == '单向双扇木质门智能门锁黑色版-1套' or Product_name_list[a] == '单向双扇木质门智能门锁金色版-1套':
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('无人收货项目对账清单明细.xlsx')

    write()

if dic1[num] == '瓦力项目对账清单明细.xlsx':
    df = pd.read_excel('瓦力项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('瓦力项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == '优博讯DT50' or Product_name_list[a] == '优博讯高拍仪（含纸膜含支架）' or Product_name_list[a] == '新北洋 Up391 移动标签打印机' or Product_name_list[a] == '优博讯高拍仪(含纸膜含支架)':
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('瓦力项目对账清单明细.xlsx')

    write()

if dic1[num] == '手表一代项目对账清单明细.xlsx':
    df = pd.read_excel('手表一代项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('手表一代项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == '41mm智能手表 电信版' or Product_name_list[a] == '41mm智能手表联通版':
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('手表一代项目对账清单明细.xlsx')

    write()

if dic1[num] == '手表二代项目对账清单明细.xlsx':
    df = pd.read_excel('手表二代项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('手表二代项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == '42':
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('手表二代项目对账清单明细.xlsx')

    write()

if dic1[num] == 'RFID智能货物管理项目对账清单明细.xlsx':
    df = pd.read_excel('RFID智能货物管理项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('RFID智能货物管理项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == 'RFID标签夹组合（280只标准装）' or Product_name_list[a] == 'RFID标签夹组合（280只无90ETC）' or Product_name_list[a] == 'RFID夹组合（280只无90ETC）-5个*56袋/箱' or Product_name_list[a] == 'RFID夹组合（280只标准装）-5个*56袋/箱' or Product_name_list[a] == '智能货物管理（PAD）-1台/箱' or Product_name_list[a] == '智能货物管理（边缘网关）-1台/箱' or Product_name_list[a] == '智能货物管理（库内基站）-1台/箱' or Product_name_list[a] == '智能货物管理（库外基站）-1台/箱' or Product_name_list[a] == '智能货物管理（门禁监测系统）-1台/箱':
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('RFID智能货物管理项目对账清单明细.xlsx')

    write()

if dic1[num] == 'PHAI项目对账清单明细.xlsx':
    df = pd.read_excel('PHAI项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('PHAI项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == '东舜时代POE供电模块 TE_POE101' or Product_name_list[a] == '明略警戒摄像机 MLC2Y1' or Product_name_list[a] == '明胜密码盘控制器 MLCM01' or Product_name_list[a] == '明胜声光报警器 MLA001_L' or Product_name_list[a] == '明胜密码盘控制器电源适配器' :
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('PHAI项目对账清单明细.xlsx')

    write()

if dic1[num] == 'ELO叫号屏项目对账清单明细.xlsx':
    df = pd.read_excel('ELO叫号屏项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('ELO叫号屏项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == 'ELO 10寸叫号一体机':
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('ELO叫号屏项目对账清单明细.xlsx')

    write()

if dic1[num] == 'IOT项目对账清单明细.xlsx':
    df = pd.read_excel('IOT项目对账清单明细.xlsx', sheet_name = 'Sheet1')
    JDE_list = df['市场/仓储'].values
    df = pd.read_excel('Bridge.xls', sheet_name = 'Sheet1')
    Name_list = df['需方法人公司名称'].values
    Name_list2 = df['需方开票单位'].values
    Name_list3 = df['店名'].values
    Name_list4 = df['上报类型'].values
    Name_list5 = df['订单名称'].values
    Name_list6 = df['JDE资产号'].values
    Product_id_list = df['JDE订单号'].values
    Product_name_num = df['设备JDE编码'].values
    Product_name_list = df['设施/设备名称'].values
    Order_num=df['采购数量'].values
    Price = df['单价（税后）'].values
    Market = df['市场'].values
    today = datetime.date.today()


    def write():
        Apend_len = len(JDE_list) + 2
        bg = op.load_workbook('IOT项目对账清单明细.xlsx')
        sheet = bg['Sheet1']
        a = 0
        while True:
            if Product_name_list[a] == 'IOT 控制箱':
                add_Name_list5.append(Name_list5[a])
                add_Market.append(Market[a])
                add_Name_list.append(Name_list[a])
                add_Name_list2.append(Name_list2[a])
                add_Product_id_list.append(Product_id_list[a])
                add_Product_name_num.append(Product_name_num[a])
                add_Product_name_list.append(Product_name_list[a])
                add_Name_list6.append(Name_list6[a])
                add_Name_list3.append(Name_list3[a])
                add_Name_list4.append(Name_list4[a])
                add_Order_num.append(Order_num[a])
                add_Price.append(Price[a])

            a = a + 1
            if a == len(Product_id_list):
                break
        print('*************************')
        print('共登录项目计数：',len(add_Name_list5))
        for b in range(0,len(add_Name_list5)):

            sheet.cell(Apend_len, 1, add_Name_list5[b])
            sheet.cell(Apend_len, 2, today)
            sheet.cell(Apend_len, 3, add_Market[b])
            sheet.cell(Apend_len, 4, add_Name_list3[b])
            sheet.cell(Apend_len, 5, add_Name_list[b])
            sheet.cell(Apend_len, 6, add_Name_list2[b])
            sheet.cell(Apend_len, 7, add_Product_id_list[b])
            sheet.cell(Apend_len, 8, add_Product_name_num [b])
            sheet.cell(Apend_len, 9, add_Product_name_list[b])
            sheet.cell(Apend_len, 10, add_Order_num[b])
            sheet.cell(Apend_len, 11, add_Price[b])
            sheet.cell(Apend_len, 12, add_Name_list6[b])
            if add_Name_list4 [b] == 'x':
                sheet.cell(Apend_len, 16, '新店')
            else:

                print('批量店名',add_Name_list3[b])

            Apend_len += 1
        bg.save('IOT项目对账清单明细.xlsx')

    write()

if dic1[num] == 'IOT-485项目对账清单明细.xlsx':
        df = pd.read_excel('IOT-485项目对账清单明细.xlsx', sheet_name='Sheet1')
        JDE_list = df['市场/仓储'].values
        df = pd.read_excel('Bridge.xls', sheet_name='Sheet1')
        Name_list = df['需方法人公司名称'].values
        Name_list2 = df['需方开票单位'].values
        Name_list3 = df['店名'].values
        Name_list4 = df['上报类型'].values
        Name_list5 = df['订单名称'].values
        Name_list6 = df['JDE资产号'].values
        Product_id_list = df['JDE订单号'].values
        Product_name_num = df['设备JDE编码'].values
        Product_name_list = df['设施/设备名称'].values
        Order_num = df['采购数量'].values
        Price = df['单价（税后）'].values
        Market = df['市场'].values
        today = datetime.date.today()


        def write():
            Apend_len = len(JDE_list) + 2
            bg = op.load_workbook('IOT-485项目对账清单明细.xlsx')
            sheet = bg['Sheet1']
            a = 0
            while True:
                if Product_name_list[a] == 'IOT 项目用 485Hub-1台':
                    add_Name_list5.append(Name_list5[a])
                    add_Market.append(Market[a])
                    add_Name_list.append(Name_list[a])
                    add_Name_list2.append(Name_list2[a])
                    add_Product_id_list.append(Product_id_list[a])
                    add_Product_name_num.append(Product_name_num[a])
                    add_Product_name_list.append(Product_name_list[a])
                    add_Name_list6.append(Name_list6[a])
                    add_Name_list3.append(Name_list3[a])
                    add_Name_list4.append(Name_list4[a])
                    add_Order_num.append(Order_num[a])
                    add_Price.append(Price[a])

                a = a + 1
                if a == len(Product_id_list):
                    break
            print('*************************')
            print('共登录项目计数：', len(add_Name_list5))
            for b in range(0, len(add_Name_list5)):

                sheet.cell(Apend_len, 1, add_Name_list5[b])
                sheet.cell(Apend_len, 2, today)
                sheet.cell(Apend_len, 3, add_Market[b])
                sheet.cell(Apend_len, 4, add_Name_list3[b])
                sheet.cell(Apend_len, 5, add_Name_list[b])
                sheet.cell(Apend_len, 6, add_Name_list2[b])
                sheet.cell(Apend_len, 7, add_Product_id_list[b])
                sheet.cell(Apend_len, 8, add_Product_name_num[b])
                sheet.cell(Apend_len, 9, add_Product_name_list[b])
                sheet.cell(Apend_len, 10, add_Order_num[b])
                sheet.cell(Apend_len, 11, add_Price[b])
                sheet.cell(Apend_len, 12, add_Name_list6[b])
                if add_Name_list4[b] == 'x':
                    sheet.cell(Apend_len, 16, '新店')
                else:

                    print('批量店名', add_Name_list3[b])

                Apend_len += 1
            bg.save('IOT-485项目对账清单明细.xlsx')


        write()
num1=input('按回车键关闭')